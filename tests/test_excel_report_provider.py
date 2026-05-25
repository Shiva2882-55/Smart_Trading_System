from pathlib import Path

from openpyxl import load_workbook

from stock_agent.domain.run_report import RunReport
from stock_agent.models import Recommendation, StockSnapshot
from stock_agent.providers.excel_report_provider import ExcelReportProvider


def _build_recommendation(
    ticker: str,
    current_price: float,
    action: str = "BUY_NOW",
    score: float = 82.0,
) -> Recommendation:
    snapshot = StockSnapshot(
        ticker=ticker,
        sector="Technology",
        company_name=f"{ticker} Limited",
        current_price=current_price,
        change_percent_3m=10.0,
        change_percent_6m=15.0,
        revenue_growth=0.12,
        earnings_growth=0.14,
        profit_margin=0.18,
        return_on_equity=0.2,
        forward_pe=22.0,
        debt_to_equity=25.0,
        trailing_eps=10.0,
        average_volume=1200000.0,
        market_cap=1000000000.0,
        annualized_volatility=0.2,
        max_drawdown_6m=0.1,
    )
    return Recommendation(
        ticker=ticker,
        company_name=snapshot.company_name,
        sector=snapshot.sector,
        score=score,
        confidence_score=78.0,
        signal="BUY",
        reasons=["Strong momentum", "Healthy growth"],
        snapshot=snapshot,
        risk_score=28.0,
        sentiment_trend=0.3,
        sector_relative_strength=2.1,
        generated_at="2026-05-26T09:15:00+05:30",
        action=action,
        action_timestamp="2026-05-26T09:15:00+05:30",
        review_timestamp="2026-05-26T15:15:00+05:30",
        entry_price=current_price,
        stop_loss=round(current_price * 0.95, 2),
        take_profit=round(current_price * 1.1, 2),
        position_size_note="Normal size",
    )


def test_excel_report_dashboard_is_first_sheet_and_contains_comparison(tmp_path: Path):
    provider = ExcelReportProvider()
    first_output = tmp_path / "stock_analysis_01.xlsx"
    second_output = tmp_path / "stock_analysis_02.xlsx"

    first_report = RunReport(
        run_id="RUN-1",
        total_requested=1,
        succeeded=1,
        status="SUCCESS",
    )
    provider.write_report(
        output_path=first_output,
        recommendations=[_build_recommendation("TCS.NS", 100.0)],
        by_sector={},
        failures=[],
        run_report=first_report,
        top_n=5,
    )

    second_report = RunReport(
        run_id="RUN-2",
        total_requested=1,
        succeeded=1,
        status="SUCCESS",
    )
    provider.write_report(
        output_path=second_output,
        recommendations=[_build_recommendation("TCS.NS", 110.0, action="WATCHLIST", score=79.0)],
        by_sector={},
        failures=[],
        run_report=second_report,
        top_n=5,
    )

    workbook = load_workbook(second_output)
    dashboard = workbook[workbook.sheetnames[0]]

    assert workbook.sheetnames[0] == "Dashboard"
    assert dashboard["A1"].value == "Stock Analysis Dashboard"

    column_a_values = [dashboard.cell(row=row, column=1).value for row in range(1, dashboard.max_row + 1)]
    assert "Previous Report Comparison" in column_a_values
    assert "Was following previous report helpful?" in column_a_values
