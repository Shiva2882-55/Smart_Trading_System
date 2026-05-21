from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd

from stock_agent.agent import StockResearchAgent
from stock_agent.config import settings


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Analyze India stocks and generate buy/hold/sell recommendations.")
    parser.add_argument("--universe", type=Path, default=settings.default_universe, help="Path to a file with one ticker per line.")
    parser.add_argument("--preset", choices=["nifty50"], help="Load a built-in India stock universe.")
    parser.add_argument("--tickers", nargs="*", help="Explicit list of tickers to analyze.")
    parser.add_argument("--top", type=int, default=settings.top_n, help="How many top-ranked stocks to display.")
    parser.add_argument("--output", type=Path, default=Path("stock_analysis.xlsx"), help="Excel output file path.")
    return parser


def _recommendations_to_dataframe(recommendations: list) -> pd.DataFrame:
    rows = []
    for rank, rec in enumerate(recommendations, start=1):
        rows.append(
            {
                "priority_rank": rank,
                "ticker": rec.ticker,
                "company_name": rec.company_name,
                "sector": rec.sector,
                "signal": rec.signal,
                "action": rec.action,
                "signal_generated_at": rec.generated_at,
                "action_timestamp": rec.action_timestamp,
                "review_timestamp": rec.review_timestamp,
                "score": rec.score,
                "confidence_score": rec.confidence_score,
                "risk_score": rec.risk_score,
                "sentiment_trend": rec.sentiment_trend,
                "sector_relative_strength": rec.sector_relative_strength,
                "current_price": rec.snapshot.current_price,
                "entry_price": rec.entry_price,
                "stop_loss": rec.stop_loss,
                "take_profit": rec.take_profit,
                "change_percent_3m": rec.snapshot.change_percent_3m,
                "change_percent_6m": rec.snapshot.change_percent_6m,
                "revenue_growth": rec.snapshot.revenue_growth,
                "earnings_growth": rec.snapshot.earnings_growth,
                "profit_margin": rec.snapshot.profit_margin,
                "return_on_equity": rec.snapshot.return_on_equity,
                "forward_pe": rec.snapshot.forward_pe,
                "debt_to_equity": rec.snapshot.debt_to_equity,
                "news_sentiment_avg": round(
                    sum(item.sentiment_score for item in rec.snapshot.news) / len(rec.snapshot.news),
                    3,
                )
                if rec.snapshot.news
                else 0.0,
                "position_size_note": rec.position_size_note,
                "top_reason_1": rec.reasons[0] if len(rec.reasons) > 0 else "",
                "top_reason_2": rec.reasons[1] if len(rec.reasons) > 1 else "",
            }
        )
    return pd.DataFrame(rows)


def _sector_leaders_to_dataframe(by_sector: dict[str, list]) -> pd.DataFrame:
    rows = []
    for sector, items in sorted(by_sector.items()):
        leader = items[0]
        rows.append(
            {
                "sector": sector,
                "leader_ticker": leader.ticker,
                "leader_company": leader.company_name,
                "signal": leader.signal,
                "action": leader.action,
                "score": leader.score,
                "confidence_score": leader.confidence_score,
                "risk_score": leader.risk_score,
                "signal_generated_at": leader.generated_at,
                "current_price": leader.snapshot.current_price,
            }
        )
    return pd.DataFrame(rows)


def _build_summary_dataframe(all_ranked_df: pd.DataFrame) -> pd.DataFrame:
    buy_count = int((all_ranked_df["signal"] == "BUY").sum())
    sell_count = int((all_ranked_df["signal"] == "SELL").sum())
    watch_count = int((all_ranked_df["action"] == "WATCHLIST").sum())

    summary_rows = [
        {"metric": "Total stocks analyzed", "value": len(all_ranked_df)},
        {"metric": "BUY signals", "value": buy_count},
        {"metric": "SELL signals", "value": sell_count},
        {"metric": "Watchlist actions", "value": watch_count},
        {"metric": "Average score", "value": round(float(all_ranked_df["score"].mean()), 2) if not all_ranked_df.empty else 0.0},
        {"metric": "Average confidence", "value": round(float(all_ranked_df["confidence_score"].mean()), 2) if not all_ranked_df.empty else 0.0},
        {"metric": "Top ticker", "value": all_ranked_df.iloc[0]["ticker"] if not all_ranked_df.empty else ""},
        {"metric": "Top action", "value": all_ranked_df.iloc[0]["action"] if not all_ranked_df.empty else ""},
    ]
    return pd.DataFrame(summary_rows)


def _build_failures_dataframe(failures: list[str]) -> pd.DataFrame:
    return pd.DataFrame({"failure": failures}) if failures else pd.DataFrame(columns=["failure"])


def _history_dir_for(output_path: Path) -> Path:
    return output_path.parent / "report_history"


def _find_previous_report(output_path: Path) -> Path | None:
    history_dir = _history_dir_for(output_path)
    if not history_dir.exists():
        return None

    reports = sorted(history_dir.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    return reports[0] if reports else None


def _load_previous_ranked_sheet(previous_report_path: Path | None) -> pd.DataFrame:
    if previous_report_path is None:
        return pd.DataFrame()
    try:
        return pd.read_excel(previous_report_path, sheet_name="All Ranked Stocks")
    except Exception:
        return pd.DataFrame()


def _safe_pct_change(current_value: float, base_value: float) -> float:
    if base_value == 0:
        return 0.0
    return round(((current_value - base_value) / base_value) * 100, 2)


def _build_comparison_dataframe(current_df: pd.DataFrame, previous_df: pd.DataFrame) -> pd.DataFrame:
    if current_df.empty or previous_df.empty:
        return pd.DataFrame(
            columns=[
                "ticker",
                "previous_priority_rank",
                "current_priority_rank",
                "rank_change",
                "previous_action",
                "current_action",
                "previous_entry_price",
                "current_price_now",
                "hypothetical_return_pct",
                "outcome_label",
                "previous_signal_generated_at",
                "current_signal_generated_at",
            ]
        )

    previous_lookup = previous_df.set_index("ticker")
    rows: list[dict[str, object]] = []

    for _, current_row in current_df.iterrows():
        ticker = current_row["ticker"]
        if ticker not in previous_lookup.index:
            continue

        previous_row = previous_lookup.loc[ticker]
        if isinstance(previous_row, pd.DataFrame):
            previous_row = previous_row.iloc[0]

        previous_action = str(previous_row.get("action", ""))
        previous_entry_price = float(previous_row.get("entry_price", current_row.get("current_price", 0.0)) or 0.0)
        current_price_now = float(current_row.get("current_price", 0.0) or 0.0)
        previous_stop_loss = float(previous_row.get("stop_loss", previous_entry_price) or previous_entry_price)
        previous_take_profit = float(previous_row.get("take_profit", previous_entry_price) or previous_entry_price)

        if previous_action in {"BUY_NOW", "BUY_ON_DIP"}:
            hypothetical_return_pct = _safe_pct_change(current_price_now, previous_entry_price)
            if current_price_now >= previous_take_profit:
                outcome_label = "Target reached or exceeded"
            elif current_price_now <= previous_stop_loss:
                outcome_label = "Below stop-loss level"
            elif hypothetical_return_pct >= 0:
                outcome_label = "Open profit if followed"
            else:
                outcome_label = "Open loss if followed"
        elif previous_action in {"SELL_NOW", "AVOID"}:
            hypothetical_return_pct = round(-_safe_pct_change(current_price_now, previous_entry_price), 2)
            if hypothetical_return_pct >= 0:
                outcome_label = "Avoiding the trade helped"
            else:
                outcome_label = "Avoiding the trade missed upside"
        else:
            hypothetical_return_pct = 0.0
            outcome_label = "No trade triggered from prior report"

        previous_rank = int(previous_row.get("priority_rank", 0) or 0)
        current_rank = int(current_row.get("priority_rank", 0) or 0)

        rows.append(
            {
                "ticker": ticker,
                "previous_priority_rank": previous_rank,
                "current_priority_rank": current_rank,
                "rank_change": previous_rank - current_rank,
                "previous_action": previous_action,
                "current_action": current_row.get("action", ""),
                "previous_entry_price": round(previous_entry_price, 2),
                "current_price_now": round(current_price_now, 2),
                "hypothetical_return_pct": hypothetical_return_pct,
                "outcome_label": outcome_label,
                "previous_signal_generated_at": previous_row.get("signal_generated_at", ""),
                "current_signal_generated_at": current_row.get("signal_generated_at", ""),
                "previous_top_reason_1": previous_row.get("top_reason_1", ""),
                "current_top_reason_1": current_row.get("top_reason_1", ""),
            }
        )

    return pd.DataFrame(rows).sort_values(by=["hypothetical_return_pct", "rank_change"], ascending=[False, False])


def _build_comparison_summary_dataframe(comparison_df: pd.DataFrame, previous_report_path: Path | None) -> pd.DataFrame:
    if comparison_df.empty:
        return pd.DataFrame(
            [
                {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
                {"metric": "Comparison rows", "value": 0},
            ]
        )

    profitable = int((comparison_df["hypothetical_return_pct"] > 0).sum())
    losing = int((comparison_df["hypothetical_return_pct"] < 0).sum())
    avg_return = round(float(comparison_df["hypothetical_return_pct"].mean()), 2)

    return pd.DataFrame(
        [
            {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
            {"metric": "Comparison rows", "value": len(comparison_df)},
            {"metric": "Positive hypothetical outcomes", "value": profitable},
            {"metric": "Negative hypothetical outcomes", "value": losing},
            {"metric": "Average hypothetical return pct", "value": avg_return},
        ]
    )


def _format_workbook(writer, all_ranked_df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = writer.book
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    blue_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 36)

    for sheet_name in [
        "All Ranked Stocks",
        "Top Recommendations",
        "Bottom Signals",
        "Best Buy Opportunities",
        "Best Sell Candidates",
        "Sector Leaders",
        "Compare With Previous",
    ]:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in sheet[1]}
        signal_col = headers.get("signal")
        action_col = headers.get("action")
        confidence_col = headers.get("confidence_score")

        for row in range(2, sheet.max_row + 1):
            if signal_col:
                signal_value = sheet.cell(row=row, column=signal_col).value
                if signal_value == "BUY":
                    sheet.cell(row=row, column=signal_col).fill = green_fill
                elif signal_value == "SELL":
                    sheet.cell(row=row, column=signal_col).fill = red_fill
                else:
                    sheet.cell(row=row, column=signal_col).fill = yellow_fill

            if action_col:
                action_value = sheet.cell(row=row, column=action_col).value
                if action_value in {"BUY_NOW", "BUY_ON_DIP"}:
                    sheet.cell(row=row, column=action_col).fill = green_fill
                elif action_value in {"SELL_NOW", "AVOID"}:
                    sheet.cell(row=row, column=action_col).fill = red_fill
                else:
                    sheet.cell(row=row, column=action_col).fill = yellow_fill

            if confidence_col:
                confidence_value = sheet.cell(row=row, column=confidence_col).value
                if isinstance(confidence_value, (int, float)) and confidence_value >= 75:
                    sheet.cell(row=row, column=confidence_col).fill = blue_fill


def _export_to_excel(
    output_path: Path,
    all_ranked_df: pd.DataFrame,
    top_df: pd.DataFrame,
    bottom_df: pd.DataFrame,
    sector_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    buy_df: pd.DataFrame,
    sell_df: pd.DataFrame,
    failures_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    comparison_summary_df: pd.DataFrame,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    history_dir = _history_dir_for(output_path)
    history_dir.mkdir(parents=True, exist_ok=True)
    archive_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_path = history_dir / f"{output_path.stem}_{archive_suffix}{output_path.suffix}"

    def _write_excel(target_path: Path) -> None:
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Report Summary", index=False)
            comparison_summary_df.to_excel(writer, sheet_name="Comparison Summary", index=False)
            all_ranked_df.to_excel(writer, sheet_name="All Ranked Stocks", index=False)
            top_df.to_excel(writer, sheet_name="Top Recommendations", index=False)
            bottom_df.to_excel(writer, sheet_name="Bottom Signals", index=False)
            buy_df.to_excel(writer, sheet_name="Best Buy Opportunities", index=False)
            sell_df.to_excel(writer, sheet_name="Best Sell Candidates", index=False)
            sector_df.to_excel(writer, sheet_name="Sector Leaders", index=False)
            comparison_df.to_excel(writer, sheet_name="Compare With Previous", index=False)
            failures_df.to_excel(writer, sheet_name="Skipped Tickers", index=False)
            _format_workbook(writer, all_ranked_df)

    try:
        _write_excel(output_path)
        _write_excel(archive_path)
        return output_path
    except ModuleNotFoundError as exc:
        if exc.name == "openpyxl":
            raise ModuleNotFoundError(
                "Excel export requires openpyxl. Run: pip install openpyxl "
                "or pip install -r requirements.txt"
            ) from exc
        raise
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        _write_excel(fallback_path)
        _write_excel(archive_path)
        print(
            f"\nRequested Excel file was locked or open. "
            f"Saved to a new file instead: {fallback_path.resolve()}"
        )
        return fallback_path


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    agent = StockResearchAgent()
    if args.tickers:
        tickers = args.tickers
    elif args.preset:
        tickers = agent.load_preset_universe(args.preset)
    else:
        tickers = agent.load_universe(args.universe)
    recommendations, by_sector = agent.analyze(tickers)

    all_ranked_df = _recommendations_to_dataframe(recommendations)
    previous_report_path = _find_previous_report(args.output)
    previous_ranked_df = _load_previous_ranked_sheet(previous_report_path)
    comparison_df = _build_comparison_dataframe(all_ranked_df, previous_ranked_df)
    comparison_summary_df = _build_comparison_summary_dataframe(comparison_df, previous_report_path)
    top_df = all_ranked_df.head(args.top).copy()
    bottom_df = all_ranked_df.tail(min(args.top, len(all_ranked_df))).copy()
    buy_df = all_ranked_df[all_ranked_df["action"].isin(["BUY_NOW", "BUY_ON_DIP"])].copy()
    sell_df = all_ranked_df[all_ranked_df["action"].isin(["SELL_NOW", "AVOID"])].copy()
    sector_df = _sector_leaders_to_dataframe(by_sector)
    summary_df = _build_summary_dataframe(all_ranked_df)
    failures_df = _build_failures_dataframe(agent.last_failures)

    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)
    pd.set_option("display.max_colwidth", 80)

    print("\nTop recommendations DataFrame")
    print("=============================")
    print(top_df)

    print("\nBottom signals DataFrame")
    print("========================")
    print(bottom_df)

    print("\nSector leaders DataFrame")
    print("========================")
    print(sector_df)

    if agent.last_failures:
        print("\nSkipped tickers")
        print("===============")
        for item in agent.last_failures:
            print(f"- {item}")

    saved_path = _export_to_excel(
        args.output,
        all_ranked_df,
        top_df,
        bottom_df,
        sector_df,
        summary_df,
        buy_df,
        sell_df,
        failures_df,
        comparison_df,
        comparison_summary_df,
    )
    print(f"\nExcel file saved to: {saved_path.resolve()}")
