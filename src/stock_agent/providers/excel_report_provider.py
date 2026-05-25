from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from time import perf_counter
from uuid import uuid4

import pandas as pd

from stock_agent.config import KOLKATA_TZ
from stock_agent.domain.run_report import RunReport
from stock_agent.providers.base import ReportProvider


logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ExcelReportProvider(ReportProvider):
    def write_report(
        self,
        output_path: Path,
        recommendations: list,
        by_sector: dict[str, list],
        failures: list[str],
        run_report: RunReport,
        top_n: int,
    ) -> Path:
        all_ranked_df = self._recommendations_to_dataframe(recommendations)
        previous_report_path = self.find_previous_report(output_path)
        previous_ranked_df = self._load_previous_ranked_sheet(previous_report_path)
        comparison_df = self._build_comparison_dataframe(all_ranked_df, previous_ranked_df)
        comparison_summary_df = self._build_comparison_summary_dataframe(comparison_df, previous_report_path)
        top_df = all_ranked_df.head(top_n).copy()
        bottom_df = all_ranked_df.tail(min(top_n, len(all_ranked_df))).copy()
        buy_df = all_ranked_df[all_ranked_df["action"].isin(["BUY_NOW", "BUY_ON_DIP"])].copy()
        sell_df = all_ranked_df[all_ranked_df["action"].isin(["SELL_NOW", "AVOID"])].copy()
        sector_df = self._sector_leaders_to_dataframe(by_sector)
        summary_df = self._build_summary_dataframe(all_ranked_df)
        failures_df = self._build_failures_dataframe(failures)
        run_summary_df = self._run_summary_to_dataframe(run_report)
        provider_issues_df = self._provider_issues_to_dataframe(run_report)
        dashboard_overview_df = self._build_dashboard_overview_dataframe(run_report, summary_df, all_ranked_df)
        dashboard_top_df = self._build_dashboard_top_recommendations_dataframe(top_df)
        dashboard_comparison_summary_df = self._build_dashboard_comparison_summary_dataframe(
            comparison_df,
            previous_report_path,
        )
        dashboard_comparison_details_df = self._build_dashboard_comparison_details_dataframe(comparison_df)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        history_dir = self.history_dir_for(output_path)
        history_dir.mkdir(parents=True, exist_ok=True)
        archive_suffix = datetime.now(KOLKATA_TZ).strftime("%Y%m%d_%H%M%S")
        archive_path = history_dir / f"{output_path.stem}_{archive_suffix}{output_path.suffix}"

        def _write_excel(target_path: Path) -> None:
            temp_path = target_path.with_name(f".{target_path.stem}.{uuid4().hex}.tmp{target_path.suffix}")
            logger.info(
                "excel_report_write_started",
                extra={"event": "excel_report_write_started", "target_path": str(target_path)},
            )
            start_time = perf_counter()
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                self._write_dashboard_sheet(
                    writer=writer,
                    overview_df=dashboard_overview_df,
                    top_df=dashboard_top_df,
                    comparison_summary_df=dashboard_comparison_summary_df,
                    comparison_details_df=dashboard_comparison_details_df,
                )
                run_summary_df.to_excel(writer, sheet_name="Run Summary", index=False)
                provider_issues_df.to_excel(writer, sheet_name="Provider Issues", index=False)
                summary_df.to_excel(writer, sheet_name="Report Summary", index=False)
                comparison_summary_df.to_excel(writer, sheet_name="Comparison Summary", index=False)
                all_ranked_df.to_excel(writer, sheet_name="All Ranked Stocks", index=False)
                top_df.to_excel(writer, sheet_name="Top Recommendations", index=False)
                bottom_df.to_excel(writer, sheet_name="Bottom Signals", index=False)
                buy_df.to_excel(writer, sheet_name="Best Buy Opportunities", index=False)
                sell_df.to_excel(writer, sheet_name="Best Sell Candidates", index=False)
                sector_df.to_excel(writer, sheet_name="Sector Leaders", index=False)
                comparison_df.to_excel(writer, sheet_name="Compare With Previous", index=False)
                failures_df.to_excel(writer, sheet_name="Skipped Tickers", index=False)
                self._format_workbook(writer)
            temp_path.replace(target_path)
            logger.info(
                "excel_report_write_success",
                extra={
                    "event": "excel_report_write_success",
                    "target_path": str(target_path),
                    "latency_ms": round((perf_counter() - start_time) * 1000, 2),
                    "status": "SUCCESS",
                },
            )

        try:
            _write_excel(output_path)
            if archive_path.resolve() != output_path.resolve():
                _write_excel(archive_path)
            return output_path
        except ModuleNotFoundError as exc:
            logger.exception(
                "excel_report_write_failed",
                extra={
                    "event": "excel_report_write_failed",
                    "target_path": str(output_path),
                    "status": "FAILED",
                    "error_type": type(exc).__name__,
                    "failure_reason": str(exc),
                },
            )
            if exc.name == "openpyxl":
                raise ModuleNotFoundError(
                    "Excel export requires openpyxl. Run: pip install openpyxl or pip install -r requirements.txt"
                ) from exc
            raise
        except PermissionError:
            timestamp = datetime.now(KOLKATA_TZ).strftime("%Y%m%d_%H%M%S")
            fallback_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
            logger.warning(
                "excel_report_write_failed",
                extra={
                    "event": "excel_report_write_failed",
                    "target_path": str(output_path),
                    "fallback_path": str(fallback_path),
                    "status": "FAILED",
                    "error_type": "PermissionError",
                    "failure_reason": "Requested Excel file was locked or open.",
                },
            )
            _write_excel(fallback_path)
            if archive_path.resolve() != fallback_path.resolve():
                _write_excel(archive_path)
            return fallback_path

    @staticmethod
    def history_dir_for(output_path: Path) -> Path:
        return output_path.parent / "report_history"

    @staticmethod
    def find_previous_report(output_path: Path) -> Path | None:
        candidates: list[Path] = []
        for report in output_path.parent.glob("*.xlsx"):
            if report.resolve() != output_path.resolve():
                candidates.append(report)

        history_dir = ExcelReportProvider.history_dir_for(output_path)
        if history_dir.exists():
            for report in history_dir.glob("*.xlsx"):
                try:
                    if report.resolve() != output_path.resolve():
                        candidates.append(report)
                except OSError:
                    continue

        reports = sorted(candidates, key=lambda path: (path.stat().st_mtime, path.name), reverse=True)
        return reports[0] if reports else None

    @staticmethod
    def _load_previous_ranked_sheet(previous_report_path: Path | None) -> pd.DataFrame:
        if previous_report_path is None:
            return pd.DataFrame()
        try:
            return pd.read_excel(previous_report_path, sheet_name="All Ranked Stocks")
        except Exception:
            return pd.DataFrame()

    @staticmethod
    def _recommendations_to_dataframe(recommendations: list) -> pd.DataFrame:
        rows = []
        for rank, rec in enumerate(recommendations, start=1):
            rows.append(
                {
                    "priority_rank": rank,
                    "ticker": rec.ticker,
                    "company_name": rec.company_name,
                    "sector": rec.sector,
                    "signal": rec.signal,
                    "action": rec.action,
                    "signal_generated_at": rec.generated_at,
                    "action_timestamp": rec.action_timestamp,
                    "review_timestamp": rec.review_timestamp,
                    "score": rec.score,
                    "confidence_score": rec.confidence_score,
                    "risk_score": rec.risk_score,
                    "sentiment_trend": rec.sentiment_trend,
                    "sector_relative_strength": rec.sector_relative_strength,
                    "current_price": rec.snapshot.current_price,
                    "entry_price": rec.entry_price,
                    "stop_loss": rec.stop_loss,
                    "take_profit": rec.take_profit,
                    "change_percent_3m": rec.snapshot.change_percent_3m,
                    "change_percent_6m": rec.snapshot.change_percent_6m,
                    "revenue_growth": rec.snapshot.revenue_growth,
                    "earnings_growth": rec.snapshot.earnings_growth,
                    "profit_margin": rec.snapshot.profit_margin,
                    "return_on_equity": rec.snapshot.return_on_equity,
                    "forward_pe": rec.snapshot.forward_pe,
                    "debt_to_equity": rec.snapshot.debt_to_equity,
                    "news_sentiment_avg": round(sum(item.sentiment_score for item in rec.snapshot.news) / len(rec.snapshot.news), 3)
                    if rec.snapshot.news
                    else 0.0,
                    "position_size_note": rec.position_size_note,
                    "top_reason_1": rec.reasons[0] if len(rec.reasons) > 0 else "",
                    "top_reason_2": rec.reasons[1] if len(rec.reasons) > 1 else "",
                }
            )
        return pd.DataFrame(rows)

    @staticmethod
    def _sector_leaders_to_dataframe(by_sector: dict[str, list]) -> pd.DataFrame:
        rows = []
        for sector, items in sorted(by_sector.items()):
            leader = items[0]
            rows.append(
                {
                    "sector": sector,
                    "leader_ticker": leader.ticker,
                    "leader_company": leader.company_name,
                    "signal": leader.signal,
                    "action": leader.action,
                    "score": leader.score,
                    "confidence_score": leader.confidence_score,
                    "risk_score": leader.risk_score,
                    "signal_generated_at": leader.generated_at,
                    "current_price": leader.snapshot.current_price,
                }
            )
        return pd.DataFrame(rows)

    @staticmethod
    def _build_summary_dataframe(all_ranked_df: pd.DataFrame) -> pd.DataFrame:
        buy_count = int((all_ranked_df["signal"] == "BUY").sum())
        sell_count = int((all_ranked_df["signal"] == "SELL").sum())
        watch_count = int((all_ranked_df["action"] == "WATCHLIST").sum())
        return pd.DataFrame(
            [
                {"metric": "Total stocks analyzed", "value": len(all_ranked_df)},
                {"metric": "BUY signals", "value": buy_count},
                {"metric": "SELL signals", "value": sell_count},
                {"metric": "Watchlist actions", "value": watch_count},
                {"metric": "Average score", "value": round(float(all_ranked_df["score"].mean()), 2) if not all_ranked_df.empty else 0.0},
                {"metric": "Average confidence", "value": round(float(all_ranked_df["confidence_score"].mean()), 2) if not all_ranked_df.empty else 0.0},
                {"metric": "Top ticker", "value": all_ranked_df.iloc[0]["ticker"] if not all_ranked_df.empty else ""},
                {"metric": "Top action", "value": all_ranked_df.iloc[0]["action"] if not all_ranked_df.empty else ""},
            ]
        )

    @staticmethod
    def _build_failures_dataframe(failures: list[str]) -> pd.DataFrame:
        return pd.DataFrame({"failure": failures}) if failures else pd.DataFrame(columns=["failure"])

    @staticmethod
    def _safe_pct_change(current_value: float, base_value: float) -> float:
        if base_value == 0:
            return 0.0
        return round(((current_value - base_value) / base_value) * 100, 2)

    def _build_comparison_dataframe(self, current_df: pd.DataFrame, previous_df: pd.DataFrame) -> pd.DataFrame:
        if current_df.empty or previous_df.empty:
            return pd.DataFrame(
                columns=[
                    "ticker",
                    "previous_priority_rank",
                    "current_priority_rank",
                    "rank_change",
                    "previous_action",
                    "current_action",
                    "previous_entry_price",
                    "current_price_now",
                    "hypothetical_return_pct",
                    "outcome_label",
                    "previous_signal_generated_at",
                    "current_signal_generated_at",
                ]
            )

        previous_lookup = previous_df.set_index("ticker")
        rows: list[dict[str, object]] = []
        for _, current_row in current_df.iterrows():
            ticker = current_row["ticker"]
            if ticker not in previous_lookup.index:
                continue
            previous_row = previous_lookup.loc[ticker]
            if isinstance(previous_row, pd.DataFrame):
                previous_row = previous_row.iloc[0]
            previous_action = str(previous_row.get("action", ""))
            previous_entry_price = float(previous_row.get("entry_price", current_row.get("current_price", 0.0)) or 0.0)
            current_price_now = float(current_row.get("current_price", 0.0) or 0.0)
            previous_stop_loss = float(previous_row.get("stop_loss", previous_entry_price) or previous_entry_price)
            previous_take_profit = float(previous_row.get("take_profit", previous_entry_price) or previous_entry_price)

            if previous_action in {"BUY_NOW", "BUY_ON_DIP"}:
                hypothetical_return_pct = self._safe_pct_change(current_price_now, previous_entry_price)
                if current_price_now >= previous_take_profit:
                    outcome_label = "Target reached or exceeded"
                elif current_price_now <= previous_stop_loss:
                    outcome_label = "Below stop-loss level"
                elif hypothetical_return_pct >= 0:
                    outcome_label = "Open profit if followed"
                else:
                    outcome_label = "Open loss if followed"
            elif previous_action in {"SELL_NOW", "AVOID"}:
                hypothetical_return_pct = round(-self._safe_pct_change(current_price_now, previous_entry_price), 2)
                if hypothetical_return_pct >= 0:
                    outcome_label = "Avoiding the trade helped"
                else:
                    outcome_label = "Avoiding the trade missed upside"
            else:
                hypothetical_return_pct = 0.0
                outcome_label = "No trade triggered from prior report"

            previous_rank = int(previous_row.get("priority_rank", 0) or 0)
            current_rank = int(current_row.get("priority_rank", 0) or 0)
            rows.append(
                {
                    "ticker": ticker,
                    "previous_priority_rank": previous_rank,
                    "current_priority_rank": current_rank,
                    "rank_change": previous_rank - current_rank,
                    "previous_action": previous_action,
                    "current_action": current_row.get("action", ""),
                    "previous_entry_price": round(previous_entry_price, 2),
                    "current_price_now": round(current_price_now, 2),
                    "hypothetical_return_pct": hypothetical_return_pct,
                    "outcome_label": outcome_label,
                    "previous_signal_generated_at": previous_row.get("signal_generated_at", ""),
                    "current_signal_generated_at": current_row.get("signal_generated_at", ""),
                    "previous_top_reason_1": previous_row.get("top_reason_1", ""),
                    "current_top_reason_1": current_row.get("top_reason_1", ""),
                }
            )

        return pd.DataFrame(rows).sort_values(by=["hypothetical_return_pct", "rank_change"], ascending=[False, False])

    @staticmethod
    def _build_comparison_summary_dataframe(comparison_df: pd.DataFrame, previous_report_path: Path | None) -> pd.DataFrame:
        if comparison_df.empty:
            return pd.DataFrame(
                [
                    {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
                    {"metric": "Comparison rows", "value": 0},
                ]
            )
        profitable = int((comparison_df["hypothetical_return_pct"] > 0).sum())
        losing = int((comparison_df["hypothetical_return_pct"] < 0).sum())
        avg_return = round(float(comparison_df["hypothetical_return_pct"].mean()), 2)
        return pd.DataFrame(
            [
                {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
                {"metric": "Comparison rows", "value": len(comparison_df)},
                {"metric": "Positive hypothetical outcomes", "value": profitable},
                {"metric": "Negative hypothetical outcomes", "value": losing},
                {"metric": "Average hypothetical return pct", "value": avg_return},
            ]
        )

    @staticmethod
    def _run_summary_to_dataframe(report: RunReport) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {"metric": "Run ID", "value": report.run_id},
                {"metric": "Status", "value": report.status},
                {"metric": "Warning Severity", "value": report.warning_severity},
                {"metric": "Total Requested", "value": report.total_requested},
                {"metric": "Succeeded", "value": report.succeeded},
                {"metric": "Failed", "value": report.failed},
                {"metric": "Skipped", "value": report.skipped},
                {"metric": "Degraded Sources", "value": ", ".join(report.degraded_sources) if report.degraded_sources else "None"},
                {"metric": "Output File", "value": report.output_file_path or ""},
            ]
        )

    @staticmethod
    def _provider_issues_to_dataframe(report: RunReport) -> pd.DataFrame:
        if not report.provider_issues:
            return pd.DataFrame(columns=["ticker", "provider", "operation", "severity", "error_type", "error_message"])
        return pd.DataFrame(
            [
                {
                    "ticker": issue.ticker,
                    "provider": issue.provider,
                    "operation": issue.operation,
                    "severity": issue.severity,
                    "error_type": issue.error_type,
                    "error_message": issue.error_message,
                }
                for issue in report.provider_issues
            ]
        )

    @staticmethod
    def _build_dashboard_overview_dataframe(
        report: RunReport,
        summary_df: pd.DataFrame,
        all_ranked_df: pd.DataFrame,
    ) -> pd.DataFrame:
        summary_lookup = {
            str(row["metric"]): row["value"]
            for _, row in summary_df.iterrows()
        }
        top_company = all_ranked_df.iloc[0]["company_name"] if not all_ranked_df.empty else ""
        return pd.DataFrame(
            [
                {"metric": "Run Status", "value": report.status},
                {"metric": "Warning Severity", "value": report.warning_severity},
                {"metric": "Total Requested", "value": report.total_requested},
                {"metric": "Succeeded", "value": report.succeeded},
                {"metric": "Failed", "value": report.failed},
                {"metric": "Skipped", "value": report.skipped},
                {"metric": "Degraded Sources", "value": ", ".join(report.degraded_sources) if report.degraded_sources else "None"},
                {"metric": "Top Stock", "value": summary_lookup.get("Top ticker", "")},
                {"metric": "Top Company", "value": top_company},
                {"metric": "Top Action", "value": summary_lookup.get("Top action", "")},
                {"metric": "Average Score", "value": summary_lookup.get("Average score", 0.0)},
                {"metric": "Average Confidence", "value": summary_lookup.get("Average confidence", 0.0)},
            ]
        )

    @staticmethod
    def _build_dashboard_top_recommendations_dataframe(top_df: pd.DataFrame) -> pd.DataFrame:
        if top_df.empty:
            return pd.DataFrame(columns=["priority_rank", "ticker", "company_name", "signal", "action", "score", "confidence_score"])
        return top_df[
            ["priority_rank", "ticker", "company_name", "signal", "action", "score", "confidence_score"]
        ].copy()

    @staticmethod
    def _build_dashboard_comparison_summary_dataframe(
        comparison_df: pd.DataFrame,
        previous_report_path: Path | None,
    ) -> pd.DataFrame:
        if comparison_df.empty:
            helpful_label = "No previous report available for comparison"
            explanation = "Run one more report later to see whether following earlier actions helped."
            return pd.DataFrame(
                [
                    {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
                    {"metric": "Stocks compared", "value": 0},
                    {"metric": "Was following previous report helpful?", "value": helpful_label},
                    {"metric": "Explanation", "value": explanation},
                ]
            )

        positive = int((comparison_df["hypothetical_return_pct"] > 0).sum())
        negative = int((comparison_df["hypothetical_return_pct"] < 0).sum())
        neutral = int((comparison_df["hypothetical_return_pct"] == 0).sum())
        average_return = round(float(comparison_df["hypothetical_return_pct"].mean()), 2)

        if average_return > 0 and positive > negative:
            helpful_label = "Yes, overall it looks helpful"
            explanation = "More previous actions helped than hurt, and the average result is positive."
        elif average_return < 0 and negative > positive:
            helpful_label = "No, overall it looks unhelpful"
            explanation = "More previous actions hurt than helped, and the average result is negative."
        else:
            helpful_label = "Mixed result"
            explanation = "Some previous actions helped and some did not, so review stock-by-stock details below."

        return pd.DataFrame(
            [
                {"metric": "Previous report linked", "value": str(previous_report_path) if previous_report_path else "No"},
                {"metric": "Stocks compared", "value": len(comparison_df)},
                {"metric": "Helpful outcomes", "value": positive},
                {"metric": "Unhelpful outcomes", "value": negative},
                {"metric": "Neutral outcomes", "value": neutral},
                {"metric": "Average hypothetical return %", "value": average_return},
                {"metric": "Was following previous report helpful?", "value": helpful_label},
                {"metric": "Explanation", "value": explanation},
            ]
        )

    @staticmethod
    def _build_dashboard_comparison_details_dataframe(comparison_df: pd.DataFrame) -> pd.DataFrame:
        if comparison_df.empty:
            return pd.DataFrame(
                columns=[
                    "ticker",
                    "previous_action",
                    "current_action",
                    "hypothetical_return_pct",
                    "outcome_label",
                ]
            )
        return comparison_df[
            [
                "ticker",
                "previous_action",
                "current_action",
                "hypothetical_return_pct",
                "outcome_label",
            ]
        ].head(10).copy()

    @staticmethod
    def _write_dashboard_sheet(
        writer,
        overview_df: pd.DataFrame,
        top_df: pd.DataFrame,
        comparison_summary_df: pd.DataFrame,
        comparison_details_df: pd.DataFrame,
    ) -> None:
        sheet_name = "Dashboard"
        overview_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        top_start_row = len(overview_df) + 7
        top_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=top_start_row)
        comparison_summary_start_row = top_start_row + len(top_df) + 6
        comparison_summary_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=comparison_summary_start_row)
        comparison_details_start_row = comparison_summary_start_row + len(comparison_summary_df) + 6
        comparison_details_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=comparison_details_start_row)

        worksheet = writer.sheets[sheet_name]
        worksheet["A1"] = "Stock Analysis Dashboard"
        worksheet["A2"] = "Overview"
        worksheet.cell(row=top_start_row, column=1, value="Top Recommendations")
        worksheet.cell(row=comparison_summary_start_row, column=1, value="Previous Report Comparison")
        worksheet.cell(row=comparison_details_start_row, column=1, value="Comparison Details")
        worksheet.sheet_view.showGridLines = False

    @staticmethod
    def _format_workbook(writer) -> None:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = writer.book
        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
        blue_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet in workbook.worksheets:
            if sheet.title != "Dashboard":
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 36)

        if "Dashboard" in workbook.sheetnames:
            dashboard = workbook["Dashboard"]
            dashboard.freeze_panes = "A4"
            dashboard["A1"].font = Font(size=16, bold=True)
            for title_cell in ("A2",):
                dashboard[title_cell].font = Font(size=12, bold=True)

            for row in range(1, dashboard.max_row + 1):
                value = dashboard.cell(row=row, column=1).value
                if value in {
                    "Overview",
                    "Top Recommendations",
                    "Previous Report Comparison",
                    "Comparison Details",
                }:
                    for col in range(1, 8):
                        dashboard.cell(row=row, column=col).fill = blue_fill
                    dashboard.cell(row=row, column=1).font = Font(bold=True)

            for row in range(1, dashboard.max_row + 1):
                first = dashboard.cell(row=row, column=1).value
                second = dashboard.cell(row=row, column=2).value
                if first == "Was following previous report helpful?":
                    if isinstance(second, str) and second.startswith("Yes"):
                        dashboard.cell(row=row, column=2).fill = green_fill
                    elif isinstance(second, str) and second.startswith("No"):
                        dashboard.cell(row=row, column=2).fill = red_fill
                    else:
                        dashboard.cell(row=row, column=2).fill = yellow_fill

        for sheet_name in [
            "All Ranked Stocks",
            "Top Recommendations",
            "Bottom Signals",
            "Best Buy Opportunities",
            "Best Sell Candidates",
            "Sector Leaders",
            "Compare With Previous",
            "Provider Issues",
        ]:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = {cell.value: cell.column for cell in sheet[1]}
            signal_col = headers.get("signal")
            action_col = headers.get("action")
            confidence_col = headers.get("confidence_score")

            for row in range(2, sheet.max_row + 1):
                if signal_col:
                    signal_value = sheet.cell(row=row, column=signal_col).value
                    if signal_value == "BUY":
                        sheet.cell(row=row, column=signal_col).fill = green_fill
                    elif signal_value == "SELL":
                        sheet.cell(row=row, column=signal_col).fill = red_fill
                    else:
                        sheet.cell(row=row, column=signal_col).fill = yellow_fill
                if action_col:
                    action_value = sheet.cell(row=row, column=action_col).value
                    if action_value in {"BUY_NOW", "BUY_ON_DIP"}:
                        sheet.cell(row=row, column=action_col).fill = green_fill
                    elif action_value in {"SELL_NOW", "AVOID"}:
                        sheet.cell(row=row, column=action_col).fill = red_fill
                    else:
                        sheet.cell(row=row, column=action_col).fill = yellow_fill
                if confidence_col:
                    confidence_value = sheet.cell(row=row, column=confidence_col).value
                    if isinstance(confidence_value, (int, float)) and confidence_value >= 75:
                        sheet.cell(row=row, column=confidence_col).fill = blue_fill
